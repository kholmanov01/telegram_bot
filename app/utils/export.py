"""Low-level Excel/PDF export builders.

These functions are pure byte producers — they take a list of duck-typed task
objects (or a stats dict) and return ``bytes`` ready to be sent as a Telegram
document. Heavy third-party imports (``openpyxl``, ``reportlab``) are placed
INSIDE the functions so that importing this module never forces those deps to
load; this keeps cold-start fast and lets ``py_compile`` validate the module
even when the libraries are not installed in the local environment.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from app.utils import dates
from app.utils.formatting import _normalize_priority, _normalize_status


def _safe_attr(obj: Any, name: str, default: Any = None) -> Any:
    """Return ``getattr(obj, name, default)`` tolerating SQLAlchemy unloaded
    attributes that raise on access."""
    try:
        return getattr(obj, name, default)
    except Exception:  # noqa: BLE001 — broad on purpose for unloaded relations
        return default


def _task_row(task: Any) -> list[str]:
    """Build a single string row for a task, suitable for both Excel and PDF."""
    task_id = _safe_attr(task, "id", "")
    title = _safe_attr(task, "title", "") or ""
    priority = _normalize_priority(_safe_attr(task, "priority", None))
    status = _normalize_status(_safe_attr(task, "status", None))

    employee = _safe_attr(task, "employee", None)
    if employee is not None:
        employee_label = (
            f"{_safe_attr(employee, 'code', '') or ''} "
            f"{_safe_attr(employee, 'full_name', '') or ''}"
        ).strip() or "—"
    else:
        employee_label = "—"

    deadline = _safe_attr(task, "deadline", None)
    deadline_str = dates.format_datetime(deadline)

    completed_at = _safe_attr(task, "completed_at", None)
    completed_str = dates.format_datetime(completed_at)

    return [
        str(task_id),
        title,
        employee_label,
        f"{priority.emoji} {priority.label}",
        status.value,
        deadline_str,
        completed_str,
    ]


def tasks_to_excel_bytes(tasks: list) -> bytes:
    """Render a list of tasks as an Excel ``.xlsx`` document in memory.

    Columns: ID, Title, Employee, Priority, Status, Deadline, Completed At.
    The header row is bold.

    :param tasks: Iterable of task-like objects.
    :returns: ``.xlsx`` file content as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vazifalar"

    headers = [
        "ID",
        "Sarlavha",
        "Xodim",
        "Ustuvorlik",
        "Holat",
        "Muddat",
        "Bajarilgan vaqt",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for task in tasks:
        sheet.append(_task_row(task))

    # Reasonable column widths.
    widths = [8, 40, 28, 14, 14, 20, 20]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def tasks_to_pdf_bytes(tasks: list) -> bytes:
    """Render a list of tasks as a PDF document in memory.

    The document has a title ("Vazifalar hisoboti"), a generation timestamp,
    and a single table with the same columns as the Excel export.

    :param tasks: Iterable of task-like objects.
    :returns: PDF file content as bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title="Vazifalar hisoboti",
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]

    elements: list[Any] = []
    elements.append(Paragraph("Vazifalar hisoboti", title_style))
    elements.append(
        Paragraph(
            f"Yaratilgan: {dates.format_datetime(dates.now_utc())}",
            normal_style,
        )
    )
    elements.append(Spacer(1, 12))

    headers = [
        "ID",
        "Sarlavha",
        "Xodim",
        "Ustuvorlik",
        "Holat",
        "Muddat",
        "Bajarilgan",
    ]
    data: list[list[Any]] = [headers]
    for task in tasks:
        row = _task_row(task)
        # Wrap long text cells in Paragraphs so they word-wrap inside cells.
        wrapped = [
            Paragraph(str(cell), normal_style) if idx in (1, 2) else str(cell)
            for idx, cell in enumerate(row)
        ]
        data.append(wrapped)

    table = Table(
        data,
        colWidths=[40, 180, 140, 70, 70, 100, 100],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    return buffer.getvalue()


def stats_to_excel_bytes(stats: dict) -> bytes:
    """Render a statistics dict as an Excel ``.xlsx`` document in memory.

    The dict is rendered as two columns: ``Ko'rsatkich`` and ``Qiymat``.

    :param stats: Mapping of statistic name to value.
    :returns: ``.xlsx`` file content as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statistika"

    headers = ["Ko'rsatkich", "Qiymat"]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    pretty: dict[str, str] = {
        "total": "Jami vazifalar",
        "pending": "Kutilmoqda",
        "completed": "Bajarilgan",
        "expired": "Muddati o'tgan",
        "archived": "Arxivlangan",
        "completed_today": "Bugun bajarilgan",
        "expired_today": "Bugun muddati o'tgan",
        "success_rate": "Muvaffaqiyat darajasi",
        "avg_completion_time": "O'rtacha bajarish vaqti",
    }

    if not stats:
        sheet.append(["—", "—"])
    else:
        for key, value in stats.items():
            label = pretty.get(str(key), str(key).replace("_", " ").title())
            display_value = value
            if key == "success_rate" and isinstance(value, (int, float)):
                display_value = f"{float(value):.1f}%"
            sheet.append([label, str(display_value) if display_value is not None else ""])

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 24
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
